
from openpyxl import load_workbook
from openpyxl import Workbook

import re
import argparse
import pathlib as pl
import shutil as sh


filename_pattern = re.compile("(Teilnahmebescheinigung.*)-(\d\d\d).pdf")

def create_cmd_line_parser():
    parser = argparse.ArgumentParser(description='Rename serial letter names ')
    parser.add_argument('filename', metavar='filename')
    parser.add_argument('directory', metavar='directory')

    return parser

def open_excel(filename: str) -> Workbook:

    wb = load_workbook(filename)
    return wb

def get_file_list(directory) -> list[str]:

    p = pl.Path(directory)
    l = list(p.glob("*.pdf"))
    print(p, l)
    matching = [f for f in l if filename_pattern.match(str(f.name))]
    return matching

def get_names (directory: str, excel_file:str) -> dict[str]:
    wb = open_excel(str(pl.Path(directory, excel_file)))
    sheet = wb.active
    m_row = sheet.max_row

    # Loop will print all values
    # of first column
    ordered_names = {}
    for i in range(2, m_row + 1):
        first_name = sheet.cell(row=i, column=2).value
        name = sheet.cell(row=i, column=3).value
        print(first_name, name)
        full_name = first_name.strip() + "_" + name.strip()
        key = "%03d" % (i - 1)
        ordered_names[key] = full_name

    return ordered_names


def do_rename(directory: str, excel_file:str) -> None:

    files = get_file_list(directory)
    print(files)

    ordered_names = get_names(directory, excel_file)
    print(ordered_names)

    for f in files:
        m = filename_pattern.match(str(f.name))
        #print(m.group(1), m.group(2))

        new_name = m.group(1) + '-' + ordered_names[m.group(2)] + '.pdf'
        full_new_name = pl.Path(f.parent, new_name)
        print("rename ", f, full_new_name)
        sh.copy(f, full_new_name)


if __name__ == "__main__":

    p = create_cmd_line_parser()
    args = p.parse_args()

    #print(args)
    full_name = "C:\\Projects\\Letter\\Test_Batch-Datei\\TFA_Teilnehmerliste_Test.xlsx"
    #wb = open_excel(full_name)
    do_rename(args.directory, args.filename)